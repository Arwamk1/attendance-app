import streamlit as st
import pandas as pd
import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font

st.set_page_config(page_title="Attendance Report Generator", layout="centered")

st.title("Attendance Report Generator")

uploaded_file = st.file_uploader("Upload your Excel attendance file", type=["xlsx"])

if uploaded_file:
    # Read file
    df = pd.read_excel(uploaded_file)
    
    # Convert Date and Time
    df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
    df['Time'] = pd.to_datetime(df['Time'], format='%H:%M:%S', errors='coerce').dt.time
    
    # Create daily key per employee
    df['DayKey'] = df['User ID'].astype(str) + '_' + df['Date'].dt.strftime('%Y-%m-%d')
    
    # Extract first Attend per day
    attend_df = df[df['Mode'] == 'Attend'].sort_values(by='Time')
    attend_df = attend_df.groupby('DayKey').first().reset_index()
    attend_df = attend_df[['DayKey', 'User ID', 'Name', 'Date', 'Time']]
    attend_df.rename(columns={'Date': 'DateAttend', 'Time': 'TimeAttend'}, inplace=True)

    # Extract last Leave per day
    leave_df = df[df['Mode'] == 'Leave'].sort_values(by='Time')
    leave_df = leave_df.groupby('DayKey').last().reset_index()
    leave_df = leave_df[['DayKey', 'User ID', 'Name', 'Date', 'Time']]
    leave_df.rename(columns={'Date': 'DateLeave', 'Time': 'TimeLeave'}, inplace=True)

    # Merge attendance and leave data
    merged_df = pd.merge(attend_df, leave_df, on='DayKey', how='outer', suffixes=('_attend', '_leave'))

    # Fill missing values
    merged_df['User ID'] = merged_df['User ID_attend'].combine_first(merged_df['User ID_leave'])
    merged_df['Name'] = merged_df['Name_attend'].combine_first(merged_df['Name_leave'])
    merged_df['DateAttend'] = merged_df['DateAttend']
    merged_df['TimeAttend'] = merged_df['TimeAttend']
    merged_df['DateLeave'] = merged_df['DateLeave']
    merged_df['TimeLeave'] = merged_df['TimeLeave']
    merged_df['Day'] = merged_df['DateAttend'].combine_first(merged_df['DateLeave']).dt.day_name()

    # Calculate monthly working hours
    def calculate_monthly_hours(row):
        if pd.notnull(row['TimeAttend']) and pd.notnull(row['TimeLeave']):
            dt_in = datetime.datetime.combine(row['DateAttend'], row['TimeAttend'])
            dt_out = datetime.datetime.combine(row['DateLeave'], row['TimeLeave'])
            delta = dt_out - dt_in
            if delta.total_seconds() > 0:
                hours, remainder = divmod(delta.total_seconds(), 3600)
                minutes, seconds = divmod(remainder, 60)
                return f"{int(hours):02}:{int(minutes):02}:{int(seconds):02}"
        return ""

    merged_df['monthly hours'] = merged_df.apply(calculate_monthly_hours, axis=1)

    # Sort by User ID and date
    merged_df['SortDate'] = merged_df['DateAttend'].combine_first(merged_df['DateLeave'])
    final_df = merged_df[['User ID', 'Name', 'Day', 'DateAttend', 'TimeAttend', 'DateLeave', 'TimeLeave', 'monthly hours', 'SortDate']]
    final_df = final_df.sort_values(by=['User ID', 'SortDate']).drop(columns=['SortDate'])

    # Create Excel file
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    bold_font = Font(bold=True)
    row_cursor = 1

    for user_id, group in final_df.groupby('User ID', sort=True):
        name = group['Name'].iloc[0]
        headers = ['User ID', 'Name', 'Day', 'DateAttend', 'TimeAttend', 'DateLeave', 'TimeLeave', 'monthly hours']
        
        # Write header
        for col_num, h in enumerate(headers, start=1):
            cell = ws.cell(row=row_cursor, column=col_num, value=h)
            cell.font = bold_font
        row_cursor += 1

        total_seconds = 0
        for _, row in group.iterrows():
            if pd.notnull(row['monthly hours']) and row['monthly hours'] != "":
                h, m, s = map(int, row['monthly hours'].split(":"))
                total_seconds += h * 3600 + m * 60 + s

            row_data = [
                row['User ID'],
                name,
                row['Day'],
                row['DateAttend'].strftime('%Y-%m-%d') if pd.notnull(row['DateAttend']) else '',
                row['TimeAttend'].strftime('%H:%M:%S') if pd.notnull(row['TimeAttend']) else '',
                row['DateLeave'].strftime('%Y-%m-%d') if pd.notnull(row['DateLeave']) else '',
                row['TimeLeave'].strftime('%H:%M:%S') if pd.notnull(row['TimeLeave']) else '',
                row['monthly hours']
            ]
            for col_num, value in enumerate(row_data, start=1):
                ws.cell(row=row_cursor, column=col_num, value=value)
            row_cursor += 1

        total_hours = round(total_seconds / 3600, 2)
        overtime = round(max(0, total_seconds - 248 * 3600) / 3600, 2)

        ws.cell(row=row_cursor, column=7, value="total hours")
        ws.cell(row=row_cursor, column=8, value=total_hours)
        row_cursor += 1
        ws.cell(row=row_cursor, column=7, value="over time")
        ws.cell(row=row_cursor, column=8, value=overtime)
        row_cursor += 1
        ws.cell(row=row_cursor, column=7, value="OT file")
        row_cursor += 2

    # Save Excel in memory
    output = BytesIO()
    wb.save(output)
    st.success("✅ Excel report is ready for download")
    st.download_button("📥 Download Report", data=output.getvalue(), file_name="Final_Attendance_Report.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # Show missing records not used
    df['UniqueKey'] = df['DayKey'] + '_' + df['Mode'] + '_' + df['Time'].astype(str)

    attend_keys = merged_df[merged_df['TimeAttend'].notnull()][['DayKey', 'TimeAttend']].copy()
    attend_keys['Mode'] = 'Attend'
    attend_keys['Time'] = attend_keys['TimeAttend'].astype(str)
    attend_keys['UniqueKey'] = attend_keys['DayKey'] + '_' + attend_keys['Mode'] + '_' + attend_keys['Time']

    leave_keys = merged_df[merged_df['TimeLeave'].notnull()][['DayKey', 'TimeLeave']].copy()
    leave_keys['Mode'] = 'Leave'
    leave_keys['Time'] = leave_keys['TimeLeave'].astype(str)
    leave_keys['UniqueKey'] = leave_keys['DayKey'] + '_' + leave_keys['Mode'] + '_' + leave_keys['Time']

    used_keys = pd.concat([attend_keys['UniqueKey'], leave_keys['UniqueKey']], ignore_index=True)
    missing_df = df[~df['UniqueKey'].isin(used_keys)]

    st.subheader("❗️ Missing Records (Not Included in Report)")
    st.dataframe(missing_df[['Date', 'Time', 'User ID', 'Name', 'Mode']])


# git add .
# git commit -m "your message"
# git push origin main
